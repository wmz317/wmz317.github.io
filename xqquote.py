import datetime
import time
import requests
import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import smtplib
import os

sCode='SH511380'
#sCode= 'TSLA'
recodercycle = 600 # 60/h  320 is enough
SendMail=1 # 为1时，发送Mail,其他值不发送
TestMode =0 # 测试模式，1时运行 ，其他值为正常应用模式
if TestMode == 1:
    recodercycle=2 #测试模式只运行2min
# 请先选择保存地址
#dirname = "D:\\staf\\QNT\\" # 本地存储
#dirname = "C:\\Users\\SR\\Desktop\\QntA\\" # 本地存储
dirname = '' # romote

def sendml(file_name):
    # smtp setup
    smtp_server = 'smtp.163.com'
    usr = 'carter317'
    ps = 'LIEJBKYFFZXYADPV'
    from_addr = 'carter317@163.com'
    to_addr = 'wmz_317@sina.com'
    #生成一个空的带附件的邮件实例
    message = MIMEMultipart()
    title = '当日'+sCode+'数据'+datetime.datetime.now().strftime('%y%m%d_%H%M') 
    message['Subject'] = title
    content = 'Dear C, 请查收今日数据'+ datetime.datetime.now().strftime('%y%m%d__%H%M')
    #将正文以text的形式插入邮件中
    message.attach(MIMEText(content, 'plain', 'utf-8'))
    #读取附件的内容
    att = MIMEText(open(file_name, 'rb').read(), 'base64', 'utf-8')
    #att["Content-Type"] = 'application/octet-stream' #不明所以
    #生成附件的名称
    att.add_header('Content-Disposition', 'attachment', filename=Header(file_name,'utf-8').encode())
    #将附件内容插入邮件中
    message.attach(att)
    
    #下面3条是防止发送失败的：加header,
    message['Subject'] = Header(title, 'utf-8')
    #可能的报错原因是因为“发件人和收件人参数没有进行定义
    message['from'] = from_addr
    message['to'] = to_addr
    
    try:
        server = smtplib.SMTP_SSL(smtp_server, 465) # 启用SSL发信, 端口一般是465
        server.login(usr, ps)
        server.sendmail(from_addr, [from_addr,to_addr], message.as_string()) # 把自己也加入收件人，防止报错
        print("邮件发送成功")
        server.quit()  # 关闭连接
    except Exception as e:
        print('出现错误：',e)
        print("邮件发送失败") 

            
# ********主程序内容**********            

def mainFun():

    def quoteData(a1_Pre):
        try: 
            a1 = session.get(url=url,headers=headers).json()
            a2 = a1['data']['quote'] #主要数据[字典]     
            return a1,a2
        except Exception as e:
            print('报错1：',e)
            a1=a1_Pre
            a1['error_code']=e #有故障时使用前一次数据，并把故障备注
            return a1, a1['data']['quote'] #有故障时返回前值，避免错误退出
       
    def quote5dang(a51_Pre):
        try:
            a51 = session.get(url=url2,headers=headers).json()
            a52 = a51['data'] #主要数据[字典]
            return a51,a52
        except Exception as e:
            print('报错2：',e)
            a51=a51_Pre
            a51['error_code']=e #有故障时使用前一次数据，并把故障备注
            return a51, a51['data'] #有故障时返回前值，避免错误退出
    
    def WrtXls(a1,a2,a51,a52,rows):
        n=0
        for v in Title_quote:
            n+=1
            ws.cell(rows+1, n, a2[v])
                
        ws.cell(rows+1,columns+1,a1['data']['market']['status'])  
        ws.cell(rows+1,columns+2,a1['data']['market']['status_id'])
        ws.cell(rows+1,columns+3,a1['data']['others']['pankou_ratio'])
        ws.cell(rows+1,columns+4,a1['error_code'])
        ws.cell(rows+1,columns+5,realtime)
        # 5dang数据写入
        m=0
        for x in title_5d:
            m+=1
            ws.cell(rows+1, columns2+m, a52[x])
        ws.cell(rows+1, columns2+m, a51['error_code'])

    # 1 配置url headers 
    url= 'https://stock.xueqiu.com/v5/stock/quote.json?symbol='+sCode+'&extend=detail'
    url2='https://stock.xueqiu.com/v5/stock/realtime/pankou.json?symbol='+sCode
    headers ={"User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36"}
    session = requests.Session()
    # 第一步,向雪球网首页发送一条请求,获取cookie
    session.get(url="https://xueqiu.com",headers=headers)
    # 第二步,获取动态加载的数据
    time.sleep(1);
    #a1 = session.get(url=url,headers=headers).json()
    #a2 = a1['data']['quote'] #主要数据
    a1_init={}
    a1,a2 = quoteData(a1_init) # 取一次数据，以获取当前HQstatus
    a51_init={}
    a51,a52= quote5dang(a51_init)
    HQstatus =a1['data']['market']['status_id']

    # 2 建立excel表格,写入表头
    nowTime = datetime.datetime.now().strftime('_%m%d_%H%M')
    file_name1 = sCode+'_Quote'+nowTime+'.xlsx'
    wb = openpyxl.Workbook() #新建工作簿（也自动生成1个工作表'Sheet'）
    ws = wb["Sheet"]
    ws.title= sCode+nowTime

    # Title首行处理，目的是不同品种兼容 且不超出自定义范围 且按顺序排布。
    # 自选的title list
    Title_quote = ['timestamp', 'current', 'premium_rate', 'percent','symbol',  'high','low','avg_price', 'acc_unit_nav','limit_up','limit_down', 'amplitude','market_capital', 'iopv', 'amount', 'chg', 'last_close', 'open','volume', 'unit_nav', 'total_shares', 'time']
    title_z7z8 = ['status','status_id','pankou_ratio','error_code','实时时间']
    title_5d= ['timestamp','bp1','bc1', 'bp2', 'bc2','bp3', 'bc3', 'bp4','bc4', 'bp5', 'bc5', 'current', 'sp1', 'sc1', 'sp2', 'sc2', 'sp3', 'sc3', 'sp4', 'sc4', 'sp5', 'sc5','buypct','sellpct','diff','ratio']
    # 网站包含的list
    title1 = list(a2.keys()) #a2.keys是dict_keys格式，需要转化
    title2 = list(set(a1['data']['market'].keys()).union(set(a1['data']['others'].keys())))
    title3 = list(a52.keys())
    # 求交集
    Title_quote =[v for v in Title_quote if v in title1]
    title_z7z8 =[v for v in title_z7z8 if v in title2]+['error_code','实时时间']
    title_5d =[v for v in title_5d if v in title3]

    # 工作表第一行名称title表头写入
    for a in range(len(Title_quote)):
        ws.cell(1,a+1,Title_quote[a])
        columns=ws.max_column # 这个列数计算一次后就不再变化
    for c in range(len(title_z7z8)):
        ws.cell(1,columns+c+1,title_z7z8[c])
    columns2=ws.max_column
    for d in range(len(title_5d)):
        ws.cell(1,columns2+d+1,title_5d[d])
    ws.cell(1,columns2+d+2,'5Dang_error_code') # 为5dang增加一个error_code用于观察数据故障
    # 保存数据到指定位置
    wb.save(dirname+file_name1) # xlsx存储
    print('表头已创建')

    # 3 循环获取数据并写入xls
    for i in range(0,recodercycle):
        if HQstatus == 5 or TestMode == 1: 
            # 进行1min（20次）的读写
            for j in range(0,20):            
                a1,a2 = quoteData(a1) # get HQ quote data
                a51,a52= quote5dang(a51) # get 5dang data    
                # 更改时间为易读格式
                timestr = time.strftime('%H:%M:%S ',time.localtime(a2['timestamp']/1000))
                a2['timestamp'] = timestr      
                # 另+1网络实时时间
                realtime = datetime.datetime.now().strftime('%H:%M:%S')
                rows = ws.max_row # 获得行数
                WrtXls(a1,a2,a51,a52,rows)
                time.sleep(3) # 3s间隔
            HQstatus =a1['data']['market']['status_id'] # 为下个循环赋值
        elif HQstatus == 7:
            print('数据时间已结束，即将完成退出')
            break  # 退出
        else:
            time.sleep(60)
            a1,a2 = quoteData(a1)
            HQstatus =a1['data']['market']['status_id'] # 为下个循环赋值
        # 定时保存-5min保存一次数据
        if i%5==0:
            wb.save(dirname+file_name1) # xlsx存储
            realtime = datetime.datetime.now().strftime('%H:%M:%S')
            print(realtime+'_数据写入完成,counter:'+str(i))

    # 4 保存(发送)数据-退出
    realtime = datetime.datetime.now().strftime('%H:%M:%S')
    ws.cell(ws.max_row+1,1,'数据保存时间'+realtime)
    wb.save(dirname+file_name1) # xlsx存储
    print('数据已保存本地')
    
    if SendMail==1:
        sendml(dirname+file_name1)
        
    success_ind=1
    return success_ind



if __name__ == '__main__':
    for i in range(0,3):
            try:
                successInd = mainFun()
            except Exception as e:
                print('第'+str(i+1)+'次,运行未成功，故障：'+ e)
                if SendMail==1:
                    file_list=os.listdir() #查找xlsx文件名
                    for x in file_list:
                        if sCode in x:
                            file_name=x
                    sendml(file_name)
            
            due_time = datetime.datetime.strptime(str(datetime.datetime.utcnow().date())+'7:00', '%Y-%m-%d%H:%M')     #注意此处是utc时间(便于在不同终端运行统一)，7点是CN15点    
            n_time = datetime.datetime.utcnow()
            if successInd == 1 and n_time > due_time:    # 如果涉及到US市场，则需要mainFun返回market市场类型来判断
                print('第'+str(i+1)+'次,运行成功，再会')
                break     # 当指示任务成功后退出循环。
        

print('任务完成！__'+ str(datetime.datetime.now()+datetime.timedelta(hours=8)))
