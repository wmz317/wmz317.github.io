# -*- coding: utf-8 -*-
"""
Created on Tue Sep 29 08:30:39 2020
@author: SR
"""

import datetime
import time
import requests
import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import smtplib
import os
from pytdx.hq import TdxHq_API
from math import ceil #,floor

sCode='SH511380'
code = sCode[-6:]
#sCode= 'TSLA'

closeTrig = 0
OpenTrig = -4

TestMode =0 # 测试模式1
recodercycle = 600 # 60/h  320 is enough
SendMail=1 # 为1时发送
if TestMode == 1: recodercycle=1 #测试模式只运行1min
# 请先选择保存地址
#dirname = "D:\\staf\\QNT\\" # 本地存储
#dirname = "C:\\Users\\SR\\Desktop\\QntA\\" # 本地存储
dirname = '' # romote

today = datetime.datetime.today().strftime('%Y%m%d')
Year = datetime.datetime.today().strftime("%Y")
mmdd = datetime.datetime.today().strftime("%m%d")
if datetime.datetime.today().weekday()>4:
    dayStep=datetime.datetime.today().weekday()-4
    lastWorkDay = datetime.datetime.today() - datetime.timedelta(days=dayStep)
    mmdd = lastWorkDay.strftime("%m%d")
    print('今天非工作日，自动取前一最近工作日')
bsetf_url = 'http://www.bosera.com/jjcp/etf/files/'+code+'/'+Year+'/'+code+mmdd+'2.ETF'

def sendml(file_name):
    # smtp setup
    smtp_server = 'smtp.163.com'
    usr = 'carter317'
    ps = 'LIEJBKYFFZXYADPV'
    from_addr = 'carter317@163.com'
    to_addr = 'wmz_317@sina.com'
    #生成一个空的带附件的邮件实例
    message = MIMEMultipart()
    title = '当日'+sCode+'数据'+datetime.datetime.now().strftime('%y%m%d_%H%M') 
    message['Subject'] = title
    content = 'Dear C, 请查收今日数据'+ datetime.datetime.now().strftime('%y%m%d__%H%M')
    #将正文以text的形式插入邮件中
    message.attach(MIMEText(content, 'plain', 'utf-8'))
    #读取附件的内容
    att = MIMEText(open(file_name, 'rb').read(), 'base64', 'utf-8')
    #att["Content-Type"] = 'application/octet-stream' #不明所以
    #生成附件的名称
    att.add_header('Content-Disposition', 'attachment', filename=Header(file_name,'utf-8').encode())
    #将附件内容插入邮件中
    message.attach(att)
    
    #下面3条是防止发送失败的：加header,
    message['Subject'] = Header(title, 'utf-8')
    #可能的报错原因是因为“发件人和收件人参数没有进行定义
    message['from'] = from_addr
    message['to'] = to_addr
    
    try:
        server = smtplib.SMTP_SSL(smtp_server, 465) # 启用SSL发信, 端口一般是465
        server.login(usr, ps)
        server.sendmail(from_addr, [from_addr,to_addr], message.as_string()) # 把自己也加入收件人，防止报错
        print("邮件发送成功")
        server.quit()  # 关闭连接
    except Exception as e:
        print('出现错误：',e)
        print("邮件发送失败") 

            
# ********主程序内容**********            

def mainFun():
    def KZZ_list(bsetf_url):
        d= requests.get(bsetf_url).text
        d=d.split('TAGTAG')
        d_cash=d[0].split('EstimateCashComponent=')[1].split('.')[0]  
        d_table=d[1].split('|                              |')
        d_table.pop() #去掉最后的EndEnd字符，不能用d3=d3.pop()，因为pop返回被删除的值
        ds=[] # kzz代码 list
        ds_rate=[] # kzz代码 对应的配置比例
        for i in range(len(d_table)):
            dd1=d_table[i].split()[0]
            ds.append(dd1)
            dd2=d_table[i].split()[2].split('|')[0]
            try:
                dd2=int(dd2)
            except:
                dd2=int(d_table[i].split()[3].split('|')[0])
            ds_rate.append(dd2)
        ds_rate_dict=dict(zip(ds,ds_rate))
        # KZZ代码对应的SH SZ抬头， 11 13开头的是SH 1; 12开头的是SZ 0
        ds_title_dict={}
        for j in ds:
            if int(j[0:2]) == 12:
                ds_title_dict[j] = 0 #SZ为0
            else:
                ds_title_dict[j] = 1
        # 转化成TDX能读懂的元祖列表
        ds_tuple=[(v,k) for k,v in ds_title_dict.items()] #列表解析-表达式
        return ds,ds_tuple,ds_rate_dict,d_cash

    # 从TDX下载price数据
    def quote_TDX(ds_tuple):
        api = TdxHq_API()
        if api.connect('119.147.212.81', 7709):
            MainCodeData = api.get_security_quotes([(1,code)])
            n= ceil(len(ds_tuple)/80)
            for i in range(n):
                if i==0:
                    quot= api.get_security_quotes(ds_tuple[i*80:(i+1)*80]) #每次最多请求80个数据？
                else:
                    quot_add= api.get_security_quotes(ds_tuple[i*80:(i+1)*80])
                    quot=quot+quot_add
    
            api.disconnect()
            return quot,MainCodeData

    # 计算iopv_tdx
    def iopv_TDX(ds_tuple):
        for k in range(3):
            try: #避免连接请求异常时退出
                quot,MainCodeData = quote_TDX(ds_tuple)
                break
            except Exception:
                print('TDX连接异常')
                time.sleep(1)
        ds_price_dict={}
        #a=quot[0]['price'] #注意KZZ quote price默认是乘了10，正好是1手
        for i in range(len(quot)):
            ds_price_dict[quot[i]['code']]= quot[i]['price']
            # 注：测试发现120004 在0918日存在price为0，当天停牌无成交导致？需要判断价格为0的错误情况,用昨收，或报价平均
            if quot[i]['price'] ==0 or quot[i]['price']<quot[i]['last_close']/10:
                ds_price_dict[quot[i]['code']]= max(quot[i]['last_close'],(quot[i]['ask1']+quot[i]['bid1'])/2)
        iopv=0 # 计算的iopv是累加值，初值为0
        for j in ds:
            iopv += ds_price_dict[j] * ds_rate_dict[j]
            iopv= round(iopv,2)
        # iopv公式？现金占比？
        zhangdie = MainCodeData[0]['price']/MainCodeData[0]['last_close'] #涨跌幅带写入
        d_cash_zd = int(d_cash)*zhangdie
        qita = 737  # 其他投资品或预留现金
        #iopv1 = (iopv + d_cash_zd )/100000
        iopv_tdx = (iopv + d_cash_zd  + qita)/100000
        iopv_tdx = round(iopv_tdx,5)
        return iopv_tdx, MainCodeData[0]['price']/10 # tdx默认ETF price*10

    def quoteData(a1_Pre):
        try: 
            a1 = session.get(url=url,headers=headers).json()
            a2 = a1['data']['quote'] #主要数据[字典]     
            return a1,a2
        except Exception as e:
            print('报错1：',e)
            a1=a1_Pre
            a1['error_code']=e #有故障时使用前一次数据，并把故障备注
            return a1, a1['data']['quote'] #有故障时返回前值，避免错误退出

    #iopv的对比选用,策略待优化
    def iopv_cal(iopv_tdx,iopv_xq):
        iopvs_diff = float(iopv_tdx)-float(iopv_xq)
        df= abs(round(iopvs_diff*1000,1))
        if df <=5: # 待定
            iopv_final= iopv_tdx
        elif df <=10:
            iopv_final= iopv_tdx*0.8+iopv_xq*0.2
        else:
            iopv_final= iopv_tdx*0.6+iopv_xq*0.4
        return iopv_final

    def kelly(dlt):
        if dlt > OpenTrig and dlt < closeTrig:
            kelly_percent_position = -1  # 不变动
        elif dlt >= closeTrig:
            kelly_percent_position = 0  # 
        elif dlt <= -20:
            kelly_percent_position = 1  # 满上
        elif dlt > -10 and dlt <= OpenTrig:
            kelly_percent_position = abs(dlt) * 0.08 - 0.1
        elif dlt > -20 and dlt <= -10:
            kelly_percent_position = abs(dlt)*0.3/7 + 0.7 - 3/7
        return kelly_percent_position

    def quote5dang(a51_Pre):
        try:
            a51 = session.get(url=url2,headers=headers).json()
            a52 = a51['data'] #主要数据[字典]
            return a51,a52
        except Exception as e:
            print('报错2：',e)
            a51=a51_Pre
            a51['error_code']=e #有故障时使用前一次数据，并把故障备注
            return a51, a51['data'] #有故障时返回前值，避免错误退出
    
    def WrtXls(a1,a2,a51,a52,rows):
        n=0
        for v in Title_quote:
            n+=1
            ws.cell(rows+1, n, a2[v])
                
        ws.cell(rows+1,columns+1,a1['data']['market']['status'])  
        ws.cell(rows+1,columns+2,a1['data']['market']['status_id'])
        ws.cell(rows+1,columns+3,a1['data']['others']['pankou_ratio'])
        ws.cell(rows+1,columns+4,a1['error_code'])
        ws.cell(rows+1,columns+5,realtime)
        # 5dang数据写入
        m=0
        for x in title_5d:
            m+=1
            ws.cell(rows+1, columns2+m, a52[x])
        ws.cell(rows+1, columns2+m, a51['error_code'])

    # 1 配置url headers 
    url= 'https://stock.xueqiu.com/v5/stock/quote.json?symbol='+sCode+'&extend=detail'
    url2='https://stock.xueqiu.com/v5/stock/realtime/pankou.json?symbol='+sCode
    headers ={"User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36"}
    session = requests.Session()
    # 第一步,向xueqiu首页发送一条请求,获取cookie
    session.get(url="https://xueqiu.com",headers=headers)
    # 第二步,获取动态加载的数据
    time.sleep(1);
    #a1 = session.get(url=url,headers=headers).json()
    #a2 = a1['data']['quote'] #主要数据
    a1_init={}
    a1,a2 = quoteData(a1_init) # 取一次数据，以获取当前HQstatus
    a51_init={}
    a51,a52= quote5dang(a51_init)
    HQstatus =a1['data']['market']['status_id']
    
    # 1.1 tdx list 初始化
    ds,ds_tuple,ds_rate_dict,d_cash= KZZ_list(bsetf_url)
    
    tdx_biaotou=['price_tdx','iopv_delta','iopv_xq','iopv_tdx','iopv_final','kelly_position_list']

    # 2 建立excel表格,写入表头
    nowTime = datetime.datetime.now().strftime('_%m%d')
    file_name1 = sCode+'分笔'+nowTime+'.xlsx'
    wb = openpyxl.Workbook() #新建工作簿（也自动生成1个工作表'Sheet'）
    ws = wb["Sheet"]
    ws.title= sCode+nowTime

    # Title首行处理，目的是不同品种兼容 且不超出自定义范围 且按顺序排布。
    # 自选的title list
    Title_quote = ['timestamp', 'current', 'premium_rate', 'percent','symbol',  'high','low','avg_price', 'acc_unit_nav','limit_up','limit_down', 'amplitude','market_capital', 'iopv', 'amount', 'chg', 'last_close', 'open','volume', 'unit_nav', 'total_shares', 'time']
    title_z7z8 = ['status','status_id','pankou_ratio','error_code','实时时间']
    title_5d= ['timestamp','bp1','bc1', 'bp2', 'bc2','bp3', 'bc3', 'bp4','bc4', 'bp5', 'bc5', 'current', 'sp1', 'sc1', 'sp2', 'sc2', 'sp3', 'sc3', 'sp4', 'sc4', 'sp5', 'sc5','buypct','sellpct','diff','ratio']
    # 网站包含的list
    title1 = list(a2.keys()) #a2.keys是dict_keys格式，需要转化
    title2 = list(set(a1['data']['market'].keys()).union(set(a1['data']['others'].keys())))
    title3 = list(a52.keys())
    # 求交集
    Title_quote =[v for v in Title_quote if v in title1]
    title_z7z8 =[v for v in title_z7z8 if v in title2]+['error_code','实时时间']
    title_5d =[v for v in title_5d if v in title3]

    # 工作表第一行名称title表头写入
    for a in range(len(Title_quote)):
        ws.cell(1,a+1,Title_quote[a])
    columns=ws.max_column # 这个列数计算一次后就不再变化
    for c in range(len(title_z7z8)):
        ws.cell(1,columns+c+1,title_z7z8[c])
    columns2=ws.max_column
    for d in range(len(title_5d)):
        ws.cell(1,columns2+d+1,title_5d[d])
    ws.cell(1,columns2+d+2,'5Dang_error_code') # 为5dang增加一个error_code用于观察数据故障
    columns3=ws.max_column
    for a in range(len(tdx_biaotou)):
        ws.cell(1,columns3+a+1,tdx_biaotou[a])
    # 保存数据到指定位置
    wb.save(dirname+file_name1) # xlsx存储
    print('表头已创建')

    # 3 循环获取数据并写入xls
    for i in range(recodercycle):
        if HQstatus == 5 or TestMode == 1: 
            # 进行1min（20次）的读写
            for j in range(0,20):            
                a1,a2 = quoteData(a1) # get HQ quote data
                a51,a52= quote5dang(a51) # get 5dang data    
                # 更改时间为易读格式
                timestr = time.strftime('%H:%M:%S ',time.localtime(a2['timestamp']/1000))
                a2['timestamp'] = timestr      
                # 另+1网络实时时间
                realtime = datetime.datetime.utcnow()+datetime.timedelta(hours=8)
                realtime=realtime.time()#去掉日期，只计时间
                
                iopv_xq= a1['data']['quote']['iopv']
                iopv_tdx,price_tdx= iopv_TDX(ds_tuple) #耗时0.5s
                iopv_final= iopv_cal(iopv_tdx,iopv_xq)
                delta_iopv= round((price_tdx-iopv_final)*1000,2)
                kelly_percent_position= kelly(delta_iopv)                
                tdx_iopv_list=[price_tdx,delta_iopv,iopv_xq,iopv_tdx,iopv_final,kelly_percent_position]
                
                rows = ws.max_row # 获得行数
                WrtXls(a1,a2,a51,a52,rows)
                for k in range(len(tdx_iopv_list)):
                    ws.cell(rows+1,columns3+k+1,tdx_iopv_list[k])

                time.sleep(2.65) # 3s间隔
            HQstatus =a1['data']['market']['status_id'] # 为下个循环赋值
        elif HQstatus == 7:
            print('数据时间已结束，即将完成退出')
            break  # 退出
        else:
            time.sleep(60)
            a1,a2 = quoteData(a1)
            HQstatus =a1['data']['market']['status_id'] # 为下个循环赋值
        # 定时保存-10min保存一次数据
        if i%10==0:
            wb.save(dirname+file_name1) # xlsx存储
            realtime = datetime.datetime.now().strftime('%H:%M:%S')
            print(realtime+'_数据写入完成,counter:'+str(i))

    # 4 保存(发送)数据-退出
    realtime = datetime.datetime.now().strftime('%H:%M:%S')
    ws.cell(ws.max_row+1,1,'数据保存时间'+realtime)
    wb.save(dirname+file_name1) # xlsx存储
    print('数据已保存本地')
    
    if SendMail==1:
        sendml(dirname+file_name1)
        
    success_ind=1
    return success_ind



if __name__ == '__main__':
    today_date = datetime.datetime.today().strftime('%m%d')
    holidays=['0101','0501','0502','0503','0504','0505'
          '1001','1002','1003','1004','1005','1006','1007','1008',
          '以下明年更新','春节','清明节','端午节','中秋节']
    for i in range(0,3):
            if today_date in holidays: #法定节假日退出
                print('今天是法定节假日，再会')
                break
            try:
                successInd = mainFun()
            except Exception as e:
                print('第'+str(i+1)+'次,运行未成功，故障：'+ e)
                if SendMail==1:
                    file_list=os.listdir() #查找xlsx文件名
                    for x in file_list:
                        if sCode in x:
                            file_name=x
                    sendml(file_name)
            
            due_time = datetime.datetime.strptime(str(datetime.datetime.utcnow().date())+'7:00', '%Y-%m-%d%H:%M')     #注意此处是utc时间(便于在不同终端运行统一)，7点是CN15点    
            n_time = datetime.datetime.utcnow()
            if successInd == 1 and n_time > due_time:    # 如果涉及到US市场，则需要mainFun返回market市场类型来判断
                print('第'+str(i+1)+'次,运行成功，再会')
                break     # 当指示任务成功后退出循环。
        

print('任务完成！__'+ str(datetime.datetime.now()+datetime.timedelta(hours=8)))
